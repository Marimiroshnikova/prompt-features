"""One Excel workbook holding the whole Phase 1 + Phase 2 dataset.

Contents follow the plan's Phase 2 grouping:

  Prompt features        the Top 30 ranked on the 10-trial label
                         (experiments/out/tentrial_top30.csv)
  Model / configuration  published specs for the 3 models, from model_specs.py
  Interaction features   context_pressure, output_pressure, recency_gap,
                         task type x model id, complexity x capability

The run grid is 280 MMLU-Pro questions x 3 Gemini models x 10 answers.
Both shapes of the same data are written: one row per question (Questions_Wide)
and one row per question x model (Runs_Long).
"""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO / "experiments"))

from evaluate import OUT  # noqa: E402
from model_specs import SPECS  # noqa: E402

DEST = OUT / "mmlu_10trial_dataset.xlsx"
N_TRIALS = 10
MAX_GEN_TOKENS = 1024
FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
GROUP_FILL = {
    "identity": PatternFill("solid", fgColor="44546A"),
    "outcome": PatternFill("solid", fgColor="806000"),
    "prompt": PatternFill("solid", fgColor="1F3864"),
    "model": PatternFill("solid", fgColor="375623"),
    "interaction": PatternFill("solid", fgColor="7B3F00"),
}

# ---- categorical entries in the Top 30: keep the raw column too --------------
CATEGORICAL = {
    "f_question_type=what": ("f_question_type", "f_question_type_is_what"),
    "f_domain_hint=legal": ("f_domain_hint", "f_domain_hint_is_legal"),
    "f_question_category=Math": ("f_question_category", "f_question_category_is_Math"),
    "f_temporal_type=range": ("f_temporal_type", "f_temporal_type_is_range"),
}

MODEL_COLS = [
    "model_family", "is_preview", "is_open_source", "has_custom_tools",
    "context_window_tokens", "knowledge_cutoff_year", "max_tokens_requested",
    "output_token_limit", "temperature", "top_p",
]
INTERACTION_COLS = [
    "f_context_pressure", "f_output_pressure", "f_recency_gap",
    "model_x_category", "f_complexity_x_capability",
]


# --------------------------------------------------------------------------- #
# assembly
# --------------------------------------------------------------------------- #

def load_top30() -> pd.DataFrame:
    return pd.read_csv(OUT / "tentrial_top30.csv")


def question_features(top30: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """The 30 ranked columns, plus the raw parent of every categorical entry."""
    feat = pd.read_csv(OUT / "mmlu_prompt_features.csv", low_memory=False)
    exam = pd.read_csv(OUT / "mmlu_exam_features.csv", low_memory=False)
    exam = exam.drop(columns=[c for c in ("q_fail", "category") if c in exam.columns])
    src = feat.merge(exam, on="question_id", how="left", suffixes=("", "_exam"))

    out = src[["question_id"]].copy()
    order: list[str] = []
    for col in top30["column"]:
        if col in CATEGORICAL:
            raw, dummy = CATEGORICAL[col]
            level = col.split("=", 1)[1]
            if raw not in out.columns:
                out[raw] = src[raw]
                order.append(raw)
            out[dummy] = (src[raw].astype(str) == level).astype(int)
            order.append(dummy)
        else:
            out[col] = src[col]
            order.append(col)
    # keep the status of anything that is not fully populated
    for col in list(order):
        st = f"{col}__status"
        if st in src.columns and src[col].isna().any():
            out[st] = src[st]
            order.append(st)
    return out, order


def base_tables() -> tuple[pd.DataFrame, pd.DataFrame]:
    res = pd.read_csv(OUT / "reduced_10trial.csv", low_memory=False)
    res["question_id"] = res["question_id"].astype(int)
    for i in range(N_TRIALS):
        res[f"it_{i + 1}"] = res["answers"].astype(str).str[i]
    res["n_fail"] = N_TRIALS - res["n_correct"]

    meta = pd.read_csv(OUT / "mmlu_prompt_features.csv", low_memory=False)
    meta = meta[["question_id", "question", "n_options", "text"]].rename(
        columns={"question": "question_stem", "text": "question_plus_options"}
    )
    return res, meta


def attach_model_and_interaction(long: pd.DataFrame) -> pd.DataFrame:
    spec = pd.DataFrame.from_dict(SPECS, orient="index")
    spec.index.name = "llm_model"
    spec = spec.reset_index()
    spec = spec[spec.llm_model.isin(long.llm_model.unique())]
    long = long.merge(spec.drop(columns=["spec_source"]), on="llm_model", how="left")

    long["top_p"] = pd.NA
    long["has_custom_tools"] = long["llm_model"].str.contains("customtools", case=False)
    long["output_token_limit"] = long["max_tokens_requested"]

    tokens = pd.to_numeric(long.get("f_context_token_count"), errors="coerce")
    window = pd.to_numeric(long["context_window_tokens"], errors="coerce")
    year = pd.to_numeric(long.get("f_year_max"), errors="coerce")
    cutoff = pd.to_numeric(long["knowledge_cutoff_year"], errors="coerce")

    long["f_context_pressure"] = tokens / window
    long["f_output_pressure"] = float(MAX_GEN_TOKENS) / float(MAX_GEN_TOKENS)
    long["f_recency_gap"] = year - cutoff
    long["model_x_category"] = long["llm_model"] + "|" + long["question_category"]
    long["f_complexity_x_capability"] = tokens * (window / 1_000_000.0)
    return long


def build() -> dict[str, pd.DataFrame]:
    top30 = load_top30()
    qfeat, feat_order = question_features(top30)
    res, meta = base_tables()

    # f_context_token_count and f_year_max drive the interaction terms; they are
    # not both in the Top 30, so pull them in explicitly.
    extra = pd.read_csv(OUT / "mmlu_prompt_features.csv", low_memory=False)
    need = [c for c in ("f_context_token_count", "f_year_max")
            if c not in qfeat.columns]
    extra = extra[["question_id", *need]]

    trial_cols = [f"it_{i}" for i in range(1, N_TRIALS + 1)]
    long = (
        res[["question_id", "question_category", "llm_model", "correct_answer",
             *trial_cols, "n_correct", "n_fail", "n_blank", "fail_rate"]]
        .merge(meta, on="question_id", how="left")
        .merge(qfeat, on="question_id", how="left")
        .merge(extra, on="question_id", how="left")
    )
    q_fail = res.groupby("question_id")["fail_rate"].mean().rename("q_fail_rate")
    long = long.merge(q_fail, on="question_id", how="left")
    long = attach_model_and_interaction(long)
    long = long.sort_values(["question_id", "llm_model"]).reset_index(drop=True)

    long_cols = (
        ["question_id", "question_category", "llm_model", "correct_answer",
         "n_options", "question_stem"]
        + trial_cols
        + ["n_correct", "n_fail", "n_blank", "fail_rate", "q_fail_rate"]
        + feat_order + MODEL_COLS + INTERACTION_COLS
    )
    long_out = long[long_cols]

    # ---- wide: one row per question ------------------------------------- #
    models = sorted(res.llm_model.unique())
    wide = meta.merge(
        res.drop_duplicates("question_id")[["question_id", "question_category",
                                            "correct_answer"]],
        on="question_id", how="left")
    for m in models:
        sub = res[res.llm_model == m].set_index("question_id")
        short = m.replace("gemini-", "")
        for i in range(1, N_TRIALS + 1):
            wide[f"{short}_t{i}"] = wide["question_id"].map(sub[f"it_{i}"])
        wide[f"{short}_n_correct"] = wide["question_id"].map(sub["n_correct"])
        wide[f"{short}_fail_rate"] = wide["question_id"].map(sub["fail_rate"])
    tot = res.groupby("question_id")["n_correct"].sum()
    wide["total_correct_of_30"] = wide["question_id"].map(tot)
    wide["q_fail_rate"] = wide["question_id"].map(q_fail)
    wide = wide.merge(qfeat, on="question_id", how="left")
    wide = wide.drop(columns=["question_plus_options"]).sort_values("question_id")

    return {"long": booleans_to_01(long_out), "wide": booleans_to_01(wide), "top30": top30, "models": models,
            "feat_order": feat_order}


# --------------------------------------------------------------------------- #
# sheets
# --------------------------------------------------------------------------- #


def booleans_to_01(df: pd.DataFrame) -> pd.DataFrame:
    """openpyxl writes a Python bool as the formula =TRUE(); 0/1 is also what a
    regression wants, so every boolean column is stored as an integer flag."""
    out = df.copy()
    for col in out.columns:
        vals = set(map(str, pd.Series(out[col]).dropna().unique()))
        if vals and vals <= {"True", "False"}:
            out[col] = (
                pd.Series(out[col])
                .map({True: 1, False: 0, "True": 1, "False": 0})
                .astype("Int64")
            )
    return out


IDENTITY_MEANING = {
    "question_id": "MMLU-Pro question id; the join key and the grouping unit for every split",
    "question_category": "MMLU-Pro subject, 20 questions each across 14 subjects",
    "llm_model": "model id used for this row's 10 answers",
    "correct_answer": "gold letter",
    "n_options": "how many lettered options the item prints",
    "question_stem": "the question text the features were computed from",
    "it_1 ... it_10": "the letter returned on each of the 10 independent calls; '-' means the reply did not parse",
    "n_correct": "answers matching the gold letter, out of 10",
    "n_fail": "10 - n_correct; unparsed replies count as wrong",
    "n_blank": "how many of the 10 replies did not parse to a single letter",
    "fail_rate": "n_fail / 10 - the label for this question x model cell",
    "q_fail_rate": "mean fail_rate across the 3 models, so 30 trials per question",
    "total_correct_of_30": "Questions_Wide only: correct answers across all 3 models",
}


def dictionary(top30: pd.DataFrame, feat_order: list[str]) -> pd.DataFrame:
    rows = [
        {"group": "Identity and outcome", "column": c, "rank": None, "direction": "",
         "what it measures": m, "spearman": None, "BH q": None, "folds (of 5)": None}
        for c, m in IDENTITY_MEANING.items()
    ]
    by_col = top30.set_index("column")
    for col in feat_order:
        if col.endswith("__status"):
            rows.append({"group": "Prompt features", "column": col, "rank": None,
                         "direction": "", "what it measures":
                         f"why {col[:-8]} is blank on that row", "spearman": None,
                         "BH q": None, "folds (of 5)": None})
            continue
        ranked = col
        if col not in by_col.index:
            hits = [k for k, (raw, dum) in CATEGORICAL.items() if col in (raw, dum)]
            ranked = hits[0] if hits else None
        r = by_col.loc[ranked] if ranked in by_col.index else None
        raw_only = ranked is not None and ranked in CATEGORICAL and col == CATEGORICAL[ranked][0]
        rows.append({
            "group": "Prompt features",
            "column": col,
            "rank": (None if r is None or raw_only else int(r["rank"])),
            "direction": "" if r is None or raw_only else r["direction"],
            "what it measures": (
                f"raw value behind the ranked flag {ranked}" if raw_only
                else (r["what_we_see"] if r is not None else "")),
            "spearman": None if r is None or raw_only else round(float(r["rho"]), 4),
            "BH q": None if r is None or raw_only else round(float(r["q"]), 3),
            "folds (of 5)": None if r is None or raw_only else int(r["folds"]),
        })

    model_meaning = {
        "model_family": "model id prefix",
        "is_preview": "the id contains 'preview'",
        "is_open_source": "Gemma yes, Gemini no",
        "has_custom_tools": "the id contains 'customtools'",
        "context_window_tokens": "published input token limit",
        "knowledge_cutoff_year": "published training cutoff year; blank for the -latest alias",
        "max_tokens_requested": "generation cap used in the run (1024, from GAIA config/inference.json)",
        "output_token_limit": "same as max_tokens_requested",
        "temperature": "sampling temperature - NOT logged by the run, blank for every row",
        "top_p": "nucleus sampling - NOT logged by the run, blank for every row",
    }
    for c, m in model_meaning.items():
        rows.append({"group": "Model / configuration features", "column": c,
                     "rank": None, "direction": "", "what it measures": m,
                     "spearman": None, "BH q": None, "folds (of 5)": None})

    inter_meaning = {
        "f_context_pressure": "prompt tokens / context window - near zero on a 1M window",
        "f_output_pressure": "expected output / max generation tokens - 1024/1024 = 1 for every row, no variance",
        "f_recency_gap": "latest year named in the question minus the model cutoff year; blank when either is missing",
        "model_x_category": "task type x model id, as a string key",
        "f_complexity_x_capability": "prompt tokens x (context window / 1e6); window is a published limit, not a quality score",
    }
    for c, m in inter_meaning.items():
        rows.append({"group": "Interaction features", "column": c, "rank": None,
                     "direction": "", "what it measures": m, "spearman": None,
                     "BH q": None, "folds (of 5)": None})
    return pd.DataFrame(rows)


def model_sheet(models: list[str]) -> pd.DataFrame:
    rows = []
    for m in models:
        s = dict(SPECS[m])
        s["llm_model"] = m
        s["questions"] = 280
        s["trials_per_question"] = N_TRIALS
        rows.append(s)
    df = booleans_to_01(pd.DataFrame(rows))
    front = ["llm_model", "model_family", "questions", "trials_per_question"]
    return df[front + [c for c in df.columns if c not in front]]


def readme(long: pd.DataFrame, wide: pd.DataFrame) -> pd.DataFrame:
    n_ans = int(long["n_correct"].sum() + long["n_fail"].sum())
    lines = [
        ("MMLU-Pro miss-prediction dataset", ""),
        ("", ""),
        ("Grid", f"{len(wide)} questions x {long.llm_model.nunique()} models x "
                 f"{N_TRIALS} answers = {n_ans:,} answers"),
        ("Source run", "results copy.csv (GAIA_STUDENT_CLUB_PROJECT eval harness)"),
        ("Label", "fail_rate = wrong answers / 10 for a question x model cell; "
                  "q_fail_rate averages the 3 models (30 trials per question)"),
        ("Unparsed replies", f"{int(long.n_blank.sum())} of {n_ans:,} "
                             "(counted as wrong; n_blank says how many per cell)"),
        ("", ""),
        ("SHEETS", ""),
        ("Questions_Wide", f"{len(wide)} rows, one per question. All 30 individual "
                           "answers as columns, per-model fail rates, prompt features."),
        ("Runs_Long", f"{len(long)} rows, one per question x model. 10 trial columns, "
                      "fail_rate, and all three feature groups."),
        ("Feature_Dictionary", "every feature column, its group, and one line on what it is."),
        ("Model_Specs", "published configuration of the 3 models, with source URLs."),
        ("Summary", "fail rate by model and by subject, computed live from Runs_Long."),
        ("", ""),
        ("FEATURE GROUPS (plan Phase 2)", ""),
        ("Prompt features", "the Top 30 ranked against the 10-trial label. Where a ranked "
                            "entry is one level of a categorical (e.g. question_type=what), "
                            "the raw column is included alongside the 0/1 flag."),
        ("Model / configuration features", "published specs only. temperature and top_p were "
                                           "not logged by the run and are blank throughout."),
        ("Interaction features", "prompt x model terms from the plan. context_pressure is "
                                 "near zero and output_pressure is constant at 1 - both are "
                                 "kept so the dictionary can say why they carry no signal."),
        ("", ""),
        ("READ BEFORE MODELLING", ""),
        ("Not significant", "0 of 139 features clear BH at q<0.05. Detection floor at n=280 "
                            "is |rho|=0.211; the largest effect present is 0.165."),
        ("Baseline to beat", "model x subject, Brier 0.1534. No feature set beats it."),
        ("Boolean flags", "stored as 1/0, not TRUE/FALSE, so they drop "
                          "straight into a regression."),
        ("Blank is not zero", "an empty feature cell means the value could not be honestly "
                              "computed; the __status column next to it says why."),
        ("Full write-up", "experiments/out/TENTRIAL_FINDINGS.md and TOP30_10TRIAL.md"),
    ]
    return pd.DataFrame(lines, columns=["", ""]).set_axis(["item", "detail"], axis=1)


def group_of(col: str, feat_order: list[str]) -> str:
    if col in MODEL_COLS:
        return "model"
    if col in INTERACTION_COLS:
        return "interaction"
    if col in feat_order:
        return "prompt"
    if col in ("n_correct", "n_fail", "n_blank", "fail_rate", "q_fail_rate",
               "total_correct_of_30") or col.endswith(("_n_correct", "_fail_rate")):
        return "outcome"
    return "identity"


def style(ws, df: pd.DataFrame, feat_order: list[str], freeze: str,
          widths: dict[str, int] | None = None) -> None:
    widths = widths or {}
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=j)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = GROUP_FILL[group_of(str(col), feat_order)]
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        letter = get_column_letter(j)
        ws.column_dimensions[letter].width = widths.get(
            str(col), min(max(len(str(col)) + 2, 9), 30))
    ws.row_dimensions[1].height = 34
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = Font(name=FONT, size=10)
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions


def main() -> None:
    d = build()
    long, wide, top30 = d["long"], d["wide"], d["top30"]
    feat_order, models = d["feat_order"], d["models"]
    fdict = dictionary(top30, feat_order)
    specs = model_sheet(models)
    rd = readme(long, wide)

    with pd.ExcelWriter(DEST, engine="openpyxl") as xl:
        rd.to_excel(xl, sheet_name="README", index=False)
        wide.to_excel(xl, sheet_name="Questions_Wide", index=False)
        long.to_excel(xl, sheet_name="Runs_Long", index=False)
        fdict.to_excel(xl, sheet_name="Feature_Dictionary", index=False)
        specs.to_excel(xl, sheet_name="Model_Specs", index=False)

        wb = xl.book
        style(wb["Questions_Wide"], wide, feat_order, "D2",
              {"question_stem": 70})
        style(wb["Runs_Long"], long, feat_order, "E2", {"question_stem": 70})
        style(wb["Feature_Dictionary"], fdict, feat_order, "A2",
              {"group": 30, "column": 34, "what it measures": 80, "direction": 24})
        style(wb["Model_Specs"], specs, feat_order, "B2", {"spec_source": 70})

        ws = wb["README"]
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 100
        for row in ws.iter_rows():
            for c in row:
                c.font = Font(name=FONT, size=10)
                c.alignment = Alignment(vertical="top", wrap_text=True)
        headings = {"MMLU-Pro miss-prediction dataset", "SHEETS",
                    "FEATURE GROUPS (plan Phase 2)", "READ BEFORE MODELLING"}
        for row in ws.iter_rows(min_col=1, max_col=1):
            if row[0].value in headings:
                for c in ws[row[0].row]:
                    c.font = Font(name=FONT, size=11, bold=True)
        ws.sheet_view.showGridLines = False

        # ---- Summary: formulas over Runs_Long so it recalculates --------- #
        sm = wb.create_sheet("Summary")
        n = len(long) + 1
        mcol = get_column_letter(long.columns.get_loc("llm_model") + 1)
        ccol = get_column_letter(long.columns.get_loc("question_category") + 1)
        fcol = get_column_letter(long.columns.get_loc("n_fail") + 1)
        kcol = get_column_letter(long.columns.get_loc("n_correct") + 1)
        rng = lambda c: f"Runs_Long!${c}$2:${c}${n}"  # noqa: E731

        sm["A1"] = "Fail rate by model"
        sm.append([])
        sm["A2"], sm["B2"], sm["C2"], sm["D2"] = "model", "wrong", "answers", "fail rate"
        for i, m in enumerate(models, start=3):
            sm[f"A{i}"] = m
            sm[f"B{i}"] = f'=SUMIF({rng(mcol)},$A{i},{rng(fcol)})'
            sm[f"C{i}"] = (f'=SUMIF({rng(mcol)},$A{i},{rng(fcol)})'
                           f'+SUMIF({rng(mcol)},$A{i},{rng(kcol)})')
            sm[f"D{i}"] = f"=IFERROR(B{i}/C{i},\"\")"
        start = 3 + len(models) + 1
        sm[f"A{start}"] = "Fail rate by subject"
        sm[f"A{start + 1}"], sm[f"B{start + 1}"] = "subject", "wrong"
        sm[f"C{start + 1}"], sm[f"D{start + 1}"] = "answers", "fail rate"
        for i, c in enumerate(sorted(long.question_category.unique()), start=start + 2):
            sm[f"A{i}"] = c
            sm[f"B{i}"] = f'=SUMIF({rng(ccol)},$A{i},{rng(fcol)})'
            sm[f"C{i}"] = (f'=SUMIF({rng(ccol)},$A{i},{rng(fcol)})'
                           f'+SUMIF({rng(ccol)},$A{i},{rng(kcol)})')
            sm[f"D{i}"] = f"=IFERROR(B{i}/C{i},\"\")"
        sm.column_dimensions["A"].width = 30
        for col in "BCD":
            sm.column_dimensions[col].width = 12
        for row in sm.iter_rows():
            for c in row:
                c.font = Font(name=FONT, size=10)
                if c.column_letter == "D":
                    c.number_format = "0.0%"
        for r in (1, 2, start, start + 1):
            for c in sm[r]:
                c.font = Font(name=FONT, size=10, bold=True)
        sm.sheet_view.showGridLines = False

    print(f"wrote {DEST}")
    print(f"  Questions_Wide {wide.shape}   Runs_Long {long.shape}")
    print(f"  feature columns: {len(feat_order)} prompt, {len(MODEL_COLS)} model, "
          f"{len(INTERACTION_COLS)} interaction")


if __name__ == "__main__":
    main()
